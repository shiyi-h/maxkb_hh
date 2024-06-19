# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import sys
from io import BytesIO
import numpy as np
from dataset.models import LongText
import uuid
import re
from collections import OrderedDict


class RAGFlowExcelParser:
    def __init__(self, **kwargs):
        self.long_text_map = {}
        self.re_pat = [[r'select\s+', r'\s+from\s+'], r'create\s+table', r'insert\s+into']
        self.re_pat = [re.compile(re_str) if isinstance(re_str, str) else [re.compile(ri) for ri in re_str] 
                       for re_str in self.re_pat]
        self.sheetname_pat = re.compile(r'^\s*sheet[0-9]*\s*$')
        self.sheet_no_value_cols = {}
        self.table_no_value_cols = {}
        self.tables = []
        self.table_max_cont_num = 20000
    
    def check_no_value_col(self, ws, sheetname):
        no_value_col = []
        for i,d in enumerate(ws.columns):
            col_mark = False
            for c in d:
                if c.value:
                    value = f"{c.value}"
                    if value.strip():
                        col_mark = True 
                        break 
            if not col_mark:
                no_value_col.append(i)
        self.sheet_no_value_cols[sheetname] = no_value_col
    
    def get_tb_head(self, table, merge_dict={}):
        head_row_num = 1
        for value in table[0]:
            pos, _ = value.split('||@#||', 1)
            if pos in merge_dict:
                merge_type = merge_dict[pos]['type']
                if merge_type=='start':
                    head_row_num = max(head_row_num, merge_dict[pos]['rows'])
        tb_head = table[:head_row_num].tolist()
        return tb_head, table[head_row_num:]
    
    def match_head_with_table(self, table, tb_head_dict, merge_dict):
        if tb_head_dict:
            tb_head = tb_head_dict['values']
            has_content_cols = tb_head_dict['cols']
            table, has_content_cols = self.rm_no_val_cols(table, has_content_cols)
            if len(has_content_cols)>len(tb_head_dict['cols']):
                _tb_head = [[] for r in tb_head]
                for cn in has_content_cols:
                    if cn in tb_head_dict['cols']:
                        for i,r in enumerate(tb_head):
                            _tb_head[i].append(r[tb_head_dict['cols'].index(cn)])
                    else:
                        for i,r in enumerate(tb_head):
                            _tb_head[i].append('||@#||')
                tb_head = _tb_head
            tb_head_dict = {'values': tb_head, 'cols': has_content_cols}
        else:
            table, has_content_cols = self.rm_no_val_cols(table, [])
            tb_head, table = self.get_tb_head(table, merge_dict)
            tb_head_dict = {'values': tb_head, 'cols': has_content_cols}
        return table, tb_head_dict
                
    def html_table(self, table, sheetname='', merge_dict={}, tb_head_dict={}):
        head_pre = '' if self.sheetname_pat.match(sheetname.lower()) else f"{sheetname}."
        table, tb_head_dict = self.match_head_with_table(table, tb_head_dict, merge_dict)
        tb = f"<table><caption>{sheetname}</caption>" if sheetname else '<table>'
        
        def _html_table(_table, _level='td', _head_pre=''):
            nonlocal tb 
            for row in _table:
                tr = "<tr>"
                for value in row:
                    pos, value = value.split('||@#||', 1)
                    if pos in merge_dict:
                        if merge_dict[pos]['type']=='belong':
                            continue 
                        rowspan = merge_dict[pos]['rows']
                        colspan = merge_dict[pos]['cols']
                        value = f"{_head_pre}{value}" if value else ''
                        span = []
                        if rowspan>1:
                            span.append(f"rowspan={rowspan}")
                        if colspan>1:
                            span.append(f"colspan={colspan}")
                        span = ' '.join(span)
                        tr += f"<{_level} {span}>{value}</{_level}>"
                    else:
                        value = f"{_head_pre}{value}" if value else ''
                        tr += f"<{_level}>{value}</{_level}>"
                tr += "</tr>"
                if tr!='<tr></tr>':
                    tb += tr 
        
        _html_table(tb_head_dict['values'], 'th', head_pre)
        _html_table(table, 'td')
        tb += "</table>\n"
        return tb_head_dict, tb
    
    def md_table(self, table, sheetname='', merge_dict={}, tb_head_dict={}):
        head_pre = '' if self.sheetname_pat.match(sheetname.lower()) else f"{sheetname}."
        tb = f'## {sheetname}\n' if sheetname else '' 
        table, tb_head_dict = self.match_head_with_table(table, tb_head_dict, merge_dict)
        
        head_v_lst = ['<br>'.join([vi.split("||@#||",1)[1] for vi in v]) for v in zip(*tb_head_dict['values'])]
        tb += '| ' + ' | '.join([head_pre+value if value else '' for value in head_v_lst]) + ' |\n'
        tb += '| ' + ' | '.join(['---' for i in range(len(head_v_lst))]) + ' |\n'
        for row in table:
            tb += '| ' + ' | '.join([value.split("||@#||",1)[1] for value in row]) + ' |\n'
        return tb_head_dict, tb
    
    def rm_no_val_cols(self, table, has_content_cols):
        table = np.array(table)
        cur_has_content_cols = []
        for i in range(table.shape[1]):
            if ''.join([v.split("||@#||",1)[1] for v in table[:,i]]).strip():
                cur_has_content_cols.append(i)
        for col in set(cur_has_content_cols)-set(has_content_cols):
            has_content_cols.append(col)
        has_content_cols.sort()
        table = table[:, has_content_cols]
        return table, has_content_cols
    
    def long_text(self, text, long_texts, fnm, sheetname):
        if text in self.long_text_map:
            id = self.long_text_map.get(text)
        else:
            self.long_text_map[text] = uuid.uuid1()
            id = self.long_text_map.get(text)
            long_texts.append(LongText(id=id, text=text, text_name=f"{fnm}--{sheetname}"))
        text_api = f'[](/api/longtext/{id})'
        return text_api
    
    def code_pattern(self, value):
        value = value.lower()
        for pat in self.re_pat:
            if isinstance(pat, list):
                if all([p.search(value) for p in pat]) and len(value)>30:
                    return True 
            else:
                if pat.search(value) and len(value)>30:
                    return True 
        return False        
    
    def get_merge_dict(self, ws, chunk_cell, fnm, sheetname, long_texts):
        merge_dict = {}
        for merged_range in ws.merged_cells.ranges:
                start_cell = f"{merged_range.start_cell.row-1}-{merged_range.start_cell.column-1}"
                merge_value = ''
                for row, col in merged_range.cells:
                    value = ws.cell(row, col).value
                    value = f"{value}" if value else ''
                    value = value.replace('\n', '<br>')
                    if (chunk_cell and len(value)>chunk_cell) or self.code_pattern(value):
                        value = self.long_text(value, long_texts, fnm, sheetname)
                    if value:
                        merge_value = value
                        break 
                rows, cols = set(), set()
                for row, col in merged_range.cells:
                    cell = f'{row-1}-{col-1}'
                    merge_dict[cell] = {"value": merge_value, 'belong': start_cell}
                    if cell==start_cell:
                        merge_dict[cell]["type"] = 'start'
                    else:
                        merge_dict[cell]['type'] = 'belong'
                    rows.add(row)
                    cols.add(col)
                merge_dict[start_cell]['rows'] = len(rows)
                merge_dict[start_cell]['cols'] = len(cols)
        new = OrderedDict()
        for d in sorted(merge_dict):
            new[d] = merge_dict[d]
        return new
    
    def update_merge_dict(self, merge_dict, rn):
        update_mark = {}
        for pos in merge_dict:
            if pos.startswith(f'{rn}-'):
                start_pos = merge_dict[pos]['belong']
                if start_pos in update_mark:
                    start_rn = int(update_mark[start_pos].split('-')[0])
                    cur_rn = int(pos.split('-')[0])
                    if cur_rn>=start_rn:
                        merge_dict[pos]['belong'] = update_mark[start_pos]
                else:
                    start_rn = int(start_pos.split('-')[0])
                    merge_dict[pos]['type'] = 'start'
                    merge_dict[pos]['rows'] = merge_dict[start_pos]['rows']-(rn-start_rn)
                    merge_dict[pos]['cols'] = merge_dict[start_pos]['cols']
                    merge_dict[pos]['belong'] = pos
                    merge_dict[start_pos]['rows']=rn-start_rn
                    update_mark[start_pos] = pos 
        return merge_dict
    
    def get_content_num(self, row, merge_dict, table_type):
        if table_type=='html':
            contents = []
            for s in row:
                if "||@#||" in s:
                    _pos, _s = s.split("||@#||")
                    if _pos in merge_dict:
                        if merge_dict[_pos]['type']=='belong':
                            continue 
                        rowspan = merge_dict[_pos]['rows']
                        colspan = merge_dict[_pos]['cols']
                        span = []
                        if rowspan>1:
                            span.append(f"rowspan={rowspan}")
                        if colspan>1:
                            span.append(f"colspan={colspan}")
                        span = ' '.join(span)
                        contents.append(f"<td {span}>{_s}</td>")
                    else:
                        contents.append('<td>'+_s+'</td>')
            content_num = len(''.join(contents))
        else:
            content_num = len(''.join(['| '+s.split("||@#||")[-1]+' |' for s in row]))
        return content_num
    
    def __call__(self, fnm, buffer=None, chunk_line=None, chunk_cell=None, table_type='html'):
        if buffer is None:
            wb = load_workbook(fnm)
        else:
            wb = load_workbook(BytesIO(buffer))
        tbs = []
        sns = []
        long_texts = []
        for sheetname in wb.sheetnames:
            sn_tb = []
            ws = wb[sheetname]
            merge_dict = self.get_merge_dict(ws, chunk_cell, fnm, sheetname, long_texts)
            rows = list(ws.rows)
            if not rows:continue
            self.check_no_value_col(ws, sheetname)
            if chunk_line:
                cur_chunk_line = chunk_line if len(rows)>=(chunk_line+5) else None
            else:
                cur_chunk_line = None
            tb = []
            tb_content_num = 0
            tb_head_dict = {}
            for rn,r in enumerate(rows):
                if (cur_chunk_line and rn%cur_chunk_line==0 and rn>0) or tb_content_num>=self.table_max_cont_num:
                    merge_dict = self.update_merge_dict(merge_dict, rn)
                    if sn_tb:
                        tbs.append('\n'.join(sn_tb))
                        sns.append(sheetname)
                        sn_tb = []
                    elif tb:
                        if table_type=='html':
                            tb_head_dict, tb = self.html_table(tb, sheetname=sheetname, merge_dict=merge_dict, 
                                                               tb_head_dict=tb_head_dict)
                        else:
                            tb_head_dict, tb = self.md_table(tb, sheetname=sheetname, merge_dict=merge_dict, 
                                                             tb_head_dict=tb_head_dict)
                        tbs.append(tb)
                        sns.append(sheetname)
                        tb = []
                        tb_content_num = 0
                tr = []
                has_value = False
                for cn, c in enumerate(r):
                    if cn in self.sheet_no_value_cols[sheetname]:
                        continue
                    pos = f"{rn}-{cn}"
                    if pos in merge_dict:
                        value = merge_dict[pos]['value']
                        if value.strip() and (not has_value):
                            has_value = True
                    elif c.value is None:
                        value = ''
                    else:
                        value = f"{c.value}" 
                        if table_type!='html':
                            value = value.replace('\n', '<br>')
                        if (chunk_cell and len(value)>chunk_cell) or self.code_pattern(value):
                            value = self.long_text(value, long_texts, fnm, sheetname)
                        if not has_value:
                            has_value = True
                    tr.append(pos+'||@#||'+value)
                if has_value:
                    tb.append(tr)
                    tb_content_num += self.get_content_num(tr, merge_dict, table_type)
                else:
                    if tb:
                        # new table
                        if table_type=='html':
                            _, tb = self.html_table(tb, sheetname=sheetname, merge_dict=merge_dict, tb_head_dict=tb_head_dict)
                        else:
                            _, tb = self.md_table(tb, sheetname=sheetname, merge_dict=merge_dict, tb_head_dict=tb_head_dict)
                        # sn_tb.append(tb)
                        tbs.append(tb)
                        sns.append(sheetname)
                        tb = []
                        tb_head_dict = {}
                        tb_content_num = 0
            if tb:
                if table_type=='html':
                    _, tb = self.html_table(tb, sheetname=sheetname, merge_dict=merge_dict, tb_head_dict=tb_head_dict)
                else:
                    _, tb = self.md_table(tb, sheetname=sheetname, merge_dict=merge_dict, tb_head_dict=tb_head_dict)
                # sn_tb.append(tb)
                tbs.append(tb)
                sns.append(sheetname)
            if sn_tb:
                tbs.append('\n'.join(sn_tb))
                sns.append(sheetname)
        return tbs, sns, long_texts

    @staticmethod
    def row_number(fnm, binary):
        if fnm.split(".")[-1].lower().find("xls") >= 0:
            wb = load_workbook(BytesIO(binary))
            total = 0
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                total += len(list(ws.rows))
                return total

        if fnm.split(".")[-1].lower() in ["csv", "txt"]:
            txt = binary.decode('utf-8', errors="ignore")
            return len(txt.split("\n"))


if __name__ == "__main__":
    psr = RAGFlowExcelParser()
    psr(sys.argv[1])
